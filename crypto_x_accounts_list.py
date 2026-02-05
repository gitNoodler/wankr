"""One-off script: build Excel list of crypto X accounts from the curated list."""
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

DATA = [
    ("@cz_binance", "8M+", "✅", "Legit", "Binance updates, market vibes. Mostly legit, SEC drama."),
    ("@VitalikButerin", "5M+", "✅", "Legit", "Ethereum co-founder. Tech, scaling, philosophy. Actual builder."),
    ("@elonmusk", "180M+", "🤷‍♂️", "Meh", "Doge pumper. Crypto memes, pumps $DOGE. Half genius, half chaos."),
    ("@CryptoCobain", "1.2M+", "🚩", "Flag", "Trader memes, TA, shitcoin calls. Pumps bags then ghosts."),
    ("@CryptoWhale", "500k+", "🤷‍♂️", "Meh", "Whale alerts, big trades. Useful but front-running vibes."),
    ("@CryptoWendyO", "300k+", "✅", "Legit", "Female trader, educational vids, TA, women-in-crypto."),
    ("@TheMoonCarl", "1M+", "🚩", "Flag", "Bitcoin maxi, price predictions. Pumps alts quietly, dumps loud."),
    ("@CryptoDonAlt", "400k+", "🤷‍♂️", "Meh", "Chart wizard, market analysis. Solid TA, shady project ties."),
    ("@CryptoBirb", "200k+", "🚩", "Flag", "Bird-themed trader, signals and memes. Botted army vibes."),
    ("@Pentosh1", "500k+", "✅", "Legit", "Trader insights, macro views. Low-key, no overt shilling."),
    ("@CryptoKaleo", "600k+", "🚩", "Flag", "Chart patterns, altcoin picks. History of hyping rugs."),
    ("@CryptoGodJohn", "150k+", "🚩", "Flag", "Meme coins, airdrop hunts. Exit liquidity farmer."),
    ("@CryptoTony__", "300k+", "🤷‍♂️", "Meh", "Futures trading, signals. Mix of good and bad calls."),
    ("@CryptoChase", "200k+", "✅", "Legit", "TA breakdowns, live streams. Educational, not much shilling."),
    ("@CryptoCapo_", "400k+", "🚩", "Flag", "Bearish takes, short calls. Rugs own community discords."),
    ("@CryptoEd_NL", "100k+", "🤷‍♂️", "Meh", "Dutch trader, wave analysis. Ties to paid groups."),
    ("@CryptoMichNL", "150k+", "✅", "Legit", "Altcoin spotter, market updates. Builder energy."),
    ("@CryptoYoda1338", "100k+", "🚩", "Flag", "Mystic vibes, predictions. LARP; dumps after visions."),
    ("@CryptoCred", "300k+", "✅", "Legit", "Trading psychology, risk management. Non-scum educator."),
    ("@CryptoRubi", "120k+", "🚩", "Flag", "NFT and Web3 takes. Pumps NFT projects that floor to zero."),
    ("@beaniemaxi", "200k+", "🚩", "Flag", "Solana shill, meme coin hunter. Exit liquidity central."),
    ("@CryptoNewton", "150k+", "🤷‍♂️", "Meh", "Physics-themed TA. Nerdy, questionable affiliations."),
    ("@AnsemBull", "300k+", "🚩", "Flag", "Bonus: Solana degen king. Pumps have left more bodies than bear market."),
]

def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Crypto X Accounts"

    headers = ["X Account", "Followers", "Rating", "Category", "Notes"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row, row_data in enumerate(DATA, 2):
        for col, value in enumerate(row_data, 1):
            ws.cell(row=row, column=col, value=value)
            ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = max(
            len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[get_column_letter(col)].width = min(
            max(ws.column_dimensions[get_column_letter(col)].width, 12), 50
        )

    out_path = "crypto_x_accounts.xlsx"
    wb.save(out_path)
    print(f"Saved: {out_path}")

if __name__ == "__main__":
    main()
